"""
Custom Ink scraper - Timed Loader (10 min per page)
Keeps original columns; appends Due Date last
Runs HEADLESS for background automation
----------------------------------------------
Requires:
  pip install selenium openpyxl requests

Credentials:
  UPDATE the USERNAME and PASSWORD variables below with your CustomInk HQ login.
"""

import os, time, logging, smtplib, traceback
import requests
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from logging.handlers import RotatingFileHandler
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import *
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook

# --------- paths ---------
SCRIPT_DIR = Path(__file__).resolve().parent
LOG_PATH = str(SCRIPT_DIR / "ci_scraper.log")
EXCEL_PATH = str(SCRIPT_DIR / "customink_orders.xlsx")

# --------- email settings ---------
# Import from email_config.py if available, otherwise use defaults
try:
    from email_config import ALERT_EMAIL, SMTP_SERVER, SMTP_PORT, SMTP_USER, SMTP_PASS
except ImportError:
    ALERT_EMAIL = "alerts@example.com"
    SMTP_SERVER = "smtp.example.com"
    SMTP_PORT = 587
    SMTP_USER = ""  # Configure in email_config.py
    SMTP_PASS = ""

def send_error_email(subject, error_message):
    """Send email notification when scraper fails."""
    if not SMTP_USER or not SMTP_PASS:
        logger.warning("[Email] SMTP credentials not configured - skipping email alert")
        return False
    
    try:
        msg = MIMEMultipart()
        msg['From'] = SMTP_USER
        msg['To'] = ALERT_EMAIL
        msg['Subject'] = f"[CI Scraper Alert] {subject}"
        
        body = f"""
CI Order Scraper encountered an error:

{error_message}

----------------------------------------
Time: {time.strftime('%Y-%m-%d %H:%M:%S')}
Computer: {os.environ.get('COMPUTERNAME', 'Unknown')}
Log file: {LOG_PATH}
----------------------------------------

This is an automated message from the CI/Ooshirts Scraper.
"""
        msg.attach(MIMEText(body, 'plain'))
        
        if SMTP_PORT == 465:
            with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
                server.login(SMTP_USER, SMTP_PASS)
                server.send_message(msg)
        else:
            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=30) as server:
                server.starttls()
                server.login(SMTP_USER, SMTP_PASS)
                server.send_message(msg)
        
        logger.info(f"[Email] Alert sent to {ALERT_EMAIL}")
        return True
    except Exception as e:
        logger.error(f"[Email] Failed to send alert: {e}")
        return False

logger = logging.getLogger("ci_scraper")
logger.setLevel(logging.INFO)
if not logger.handlers:
    fh = RotatingFileHandler(LOG_PATH, maxBytes=1_000_000, backupCount=2, encoding="utf-8")
    fh.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
    logger.addHandler(fh)
    sh = logging.StreamHandler()
    sh.setFormatter(logging.Formatter("%(message)s"))
    logger.addHandler(sh)

# --------- credentials ---------
# UPDATE THESE with your CustomInk HQ login credentials
USERNAME = "your-email@example.com"  # <-- UPDATE THIS
PASSWORD = "your-password"           # <-- UPDATE THIS
LOGIN_URL = "https://hq.customink.com/users/sign_in"

# --------- URLs ---------
# These filter URLs are specific to Promos Ink vendor IDs
# You may need to update these for your vendors
BASE_FILTER = (
    "utf8=%E2%9C%93"
    "&filter%5Bvendor_id%5D%5B%5D="
    "&filter%5Bvendor_id%5D%5B%5D=3324"
    "&filter%5Bvendor_id%5D%5B%5D=3403"
    "&filter%5Bvendor_id%5D%5B%5D=6030"
    "&filter%5Bvendor_id%5D%5B%5D=6031"
    "&filter%5Bvendor_id%5D%5B%5D=6032"
    "&filter%5Bvendor_id%5D%5B%5D=6233"
    "&filter%5Bvendor_id%5D%5B%5D=8855"
    "&filter%5Bvendor_id%5D%5B%5D=8895"
    "&filter%5Bvendor_id%5D%5B%5D=8935"
    "&filter%5Bvendor_id%5D%5B%5D=9215"
    "&filter%5Bvendor_id%5D%5B%5D=9235"
    "&filter%5Bvendor_id%5D%5B%5D=9435"
    "&filter%5Bvendor_id%5D%5B%5D=9515"
    "&filter%5Bvendor_id%5D%5B%5D=9576"
    "&filter%5Bvendor_id%5D%5B%5D=9595"
    "&filter%5Bvendor_id%5D%5B%5D=9715"
    "&multiselect=3324&multiselect=3403&multiselect=6030&multiselect=6031"
    "&multiselect=6032&multiselect=6233&multiselect=8855&multiselect=8895"
    "&multiselect=8935&multiselect=9215&multiselect=9235&multiselect=9435"
    "&multiselect=9515&multiselect=9576&multiselect=9595&multiselect=9715"
    "&filter%5Btypes%5D%5B%5D="
    "&filter%5Border_ship%5D=&filter%5Bcustomink_order_id%5D="
)

STANDARD_ORDERS_URL = f"https://hq.customink.com/orders?{BASE_FILTER}"
PENDING_ORDERS_URL  = f"https://hq.customink.com/orders?{BASE_FILTER}&filter%5Btypes%5D%5B%5D=pending&filter%5Btypes%5D%5B%5D="

# --------- cloud sync ---------
def upload_to_supply_chain(file_path):
    """
    Upload CustomInk orders to Promos Ink Supply Chain cloud.
    This allows receiving stations to look up package info without 
    needing direct access to the network share.
    """
    API_URL = "https://supplychain.promosinkwall-e.com/api/manifests"
    API_KEY = "promos-ink-2024"
    
    if not os.path.exists(file_path):
        logger.warning(f"[Cloud Sync] File not found: {file_path}")
        return False
    
    try:
        file_size = os.path.getsize(file_path)
        logger.info(f"[Cloud Sync] Uploading {file_path} ({file_size:,} bytes)...")
        
        with open(file_path, 'rb') as f:
            response = requests.post(
                API_URL,
                files={'file': (os.path.basename(file_path), f)},
                data={'type': 'customink'},
                headers={'x-api-key': API_KEY},
                timeout=60
            )
        
        if response.status_code == 200:
            logger.info("[Cloud Sync] Upload successful")
            return True
        else:
            logger.warning(f"[Cloud Sync] Upload failed ({response.status_code}) - local file still saved")
            return False
            
    except Exception as e:
        logger.warning(f"[Cloud Sync] Upload error: {e} - local file still saved")
        return False

# --------- helpers ---------
def wait_for(drv, by, val, t=30):
    return WebDriverWait(drv, t).until(EC.visibility_of_element_located((by, val)))

def login(drv):
    logger.info("Logging in...")
    drv.get(LOGIN_URL)
    wait_for(drv, By.NAME, "user[email]")
    drv.find_element(By.NAME, "user[email]").send_keys(USERNAME)
    drv.find_element(By.NAME, "user[password]").send_keys(PASSWORD)
    drv.find_element(By.NAME, "commit").click()
    wait_for(drv, By.LINK_TEXT, "Orders")
    logger.info("[OK] Logged in")

def _rows(drv):
    return drv.find_elements(By.XPATH, "//tbody/tr[@data-order-id]")

def _group_due_for_row(row):
    try:
        group = row.find_element(By.XPATH, "ancestor::tbody[@data-order-group][1]")
        h4 = group.find_element(By.XPATH, ".//tr[th[@class='day']]//h4")
        return h4.text.strip()  # e.g., "Thu, Aug 07"
    except Exception:
        return ""

def _more_button(drv):
    try:
        return drv.find_element(By.XPATH, "//tfoot//a[contains(.,'Scroll Down for More Orders')]")
    except Exception:
        return None

def timed_loader(drv, duration_sec=300):
    """
    For the current page, keep loading more for exactly duration_sec seconds.
    Strategy each loop:
      - scroll to bottom
      - click 'More' if visible (normal click, fallback js click)
      - send END a few times
      - brief sleeps to let rows render
    Logs row count whenever it grows.
    """
    start = time.time()
    last_count = -1
    rounds = 0
    body = None
    try:
        body = drv.find_element(By.TAG_NAME, "body")
    except Exception:
        pass

    while time.time() - start < duration_sec:
        rounds += 1

        # Scroll down hard
        try:
            drv.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        except Exception:
            pass
        if body:
            try:
                for _ in range(3):
                    body.send_keys(Keys.END); time.sleep(0.15)
            except Exception:
                pass

        # Click the More button if present
        btn = _more_button(drv)
        if btn and btn.is_displayed() and btn.is_enabled():
            try:
                drv.execute_script("arguments[0].scrollIntoView({block:'end'});", btn)
                btn.click()
                logger.info("    [loader] clicked More (normal)")
                time.sleep(0.6)
            except Exception:
                try:
                    drv.execute_script("arguments[0].click();", btn)
                    logger.info("    [loader] clicked More (JS)")
                    time.sleep(0.6)
                except Exception:
                    pass

        # Allow content to load
        time.sleep(0.6)
        # If network/render lag, poll briefly for growth
        end_wait = time.time() + 3.0
        while time.time() < end_wait:
            c2 = len(_rows(drv))
            if c2 > last_count:
                break
            time.sleep(0.25)

        count = len(_rows(drv))
        if count != last_count:
            logger.info(f"    rows: {count} (round {rounds})")
            last_count = count

def extract(drv, pending_mode=False):
    data = []
    for r in _rows(drv):
        t = r.find_elements(By.TAG_NAME, "td")
        if len(t) < 10:
            continue

        due_text = _group_due_for_row(r)

        status_text = t[1].text.strip()
        flags_cell = ""
        try:
            flags_cell = t[2].text.strip()
        except Exception:
            pass
        if pending_mode:
            if "hold" in flags_cell.lower():
                status_text = f"{status_text} (On Hold)"
            else:
                status_text = f"{status_text} (Pipelined)"
        else:
            if "hold" in flags_cell.lower():
                status_text = f"{status_text} (On Hold)"

        # KEEP ORIGINAL 7 columns, then append Due Date LAST
        row_out = [
            t[0].text.strip(),  # Order #
            status_text,        # Status
            t[4].text.strip(),  # Vendor
            t[6].text.strip(),  # Units
            t[7].text.strip(),  # Screens
            t[8].text.strip(),  # Shipper
            t[9].text.strip(),  # Check-In
            due_text            # Due Date
        ]
        # Prefer link text for order #
        try:
            link = t[0].find_element(By.TAG_NAME, "a")
            if link.text.strip():
                row_out[0] = link.text.strip()
        except Exception:
            pass

        data.append(row_out)
    return data

def save_xlsx(path, rows):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        header = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column+1)]
        if not header:
            ws.append(["Order #", "Status", "Vendor", "Units", "Screens", "Shipper", "Check-In", "Due Date"])
        elif "Due Date" not in [str(h or "") for h in header]:
            ws.cell(row=1, column=ws.max_column+1, value="Due Date")
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Order #", "Status", "Vendor", "Units", "Screens", "Shipper", "Check-In", "Due Date"])

    existing = {str(ws.cell(row=i, column=1).value).strip() for i in range(2, ws.max_row+1)}
    added = 0
    for r in rows:
        if not r or not r[0]:
            continue
        if str(r[0]).strip() not in existing:
            ws.append(r); added += 1
    wb.save(path)
    logger.info(f"[+] Added {added} rows to {path}")

def scrape_timed(drv, url, label, minutes=5):
    logger.info(f"\n=== {label} ===")
    drv.get(url)
    wait_for(drv, By.XPATH, "//table[@id='order_details_span_12']")
    # Load for exactly N minutes
    timed_loader(drv, duration_sec=minutes*60)
    pending_mode = ("pending" in url.lower())
    return extract(drv, pending_mode=pending_mode)

# --------- main ---------
def main():
    # HEADLESS MODE - runs in background without visible browser
    opts = Options()
    opts.add_argument("--headless=new")  # New headless mode (Chrome 109+)
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    
    logger.info("=" * 50)
    logger.info("CI SCRAPER STARTED (HEADLESS MODE)")
    logger.info("=" * 50)

    drv = webdriver.Chrome(options=opts, service=Service())
    try:
        login(drv)
        std  = scrape_timed(drv, STANDARD_ORDERS_URL, "Standard Orders (10 min)", minutes=10)
        pend = scrape_timed(drv, PENDING_ORDERS_URL,  "Pending Orders (10 min)", minutes=10)
        combined = {r[0]: r for r in std + pend}.values()
        save_xlsx(EXCEL_PATH, list(combined))
        logger.info("All done.")
        
        # Upload to Promos Ink Supply Chain cloud
        upload_to_supply_chain(EXCEL_PATH)
    finally:
        try: drv.quit()
        except Exception: pass

if __name__ == "__main__":
    try:
        main()
        logger.info("[OK] CI Scraper completed successfully")
    except Exception as e:
        error_msg = traceback.format_exc()
        logger.exception("Fatal error")
        # Send email alert
        send_error_email("Scraper Failed", error_msg)
        print("\n[!] An error occurred. See ci_scraper.log for details.")
