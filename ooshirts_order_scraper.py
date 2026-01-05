"""
Ooshirts Order Scraper - HEADLESS MODE
Runs in background for scheduled automation
----------------------------------------------
Requires:
  pip install selenium openpyxl

Credentials:
  UPDATE the OO_ACCOUNTS list below with your Ooshirts login credentials.
"""

import os, re, time, logging, traceback, smtplib
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from logging.handlers import RotatingFileHandler
from pathlib import Path

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import *
from openpyxl import Workbook, load_workbook

# Optional clipboard lib (fallback not required now)
try:
    import pyperclip
except Exception:
    pyperclip = None

SCRIPT_DIR = Path(__file__).resolve().parent
LOG_PATH   = SCRIPT_DIR / "ooshirts_scraper.log"
EXCEL_PATH = SCRIPT_DIR / "customink_orders.xlsx"

# --------- email settings ---------
# Import from email_config.py if available, otherwise use defaults
try:
    from email_config import ALERT_EMAIL, SMTP_SERVER, SMTP_PORT, SMTP_USER, SMTP_PASS
except ImportError:
    ALERT_EMAIL = "alerts@example.com"
    SMTP_SERVER = "smtp.example.com"
    SMTP_PORT = 587
    SMTP_USER = ""  # Configure in email_config.py
    SMTP_PASS = ""

def send_error_email(subject, error_message):
    """Send email notification when scraper fails."""
    if not SMTP_USER or not SMTP_PASS:
        logger.warning("[Email] SMTP credentials not configured - skipping email alert")
        return False
    
    try:
        msg = MIMEMultipart()
        msg['From'] = SMTP_USER
        msg['To'] = ALERT_EMAIL
        msg['Subject'] = f"[Ooshirts Scraper Alert] {subject}"
        
        body = f"""
Ooshirts Order Scraper encountered an error:

{error_message}

----------------------------------------
Time: {time.strftime('%Y-%m-%d %H:%M:%S')}
Computer: {os.environ.get('COMPUTERNAME', 'Unknown')}
Log file: {LOG_PATH}
----------------------------------------

This is an automated message from the CI/Ooshirts Scraper.
"""
        msg.attach(MIMEText(body, 'plain'))
        
        if SMTP_PORT == 465:
            with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
                server.login(SMTP_USER, SMTP_PASS)
                server.send_message(msg)
        else:
            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=30) as server:
                server.starttls()
                server.login(SMTP_USER, SMTP_PASS)
                server.send_message(msg)
        
        logger.info(f"[Email] Alert sent to {ALERT_EMAIL}")
        return True
    except Exception as e:
        logger.error(f"[Email] Failed to send alert: {e}")
        return False

# --------- Ooshirts config ---------
# UPDATE THESE with your Ooshirts login credentials
OO_VIEWS = [
    "https://www.ooshirts.com/?module=printer&action=orders&view=BlanksNotReady",
    "https://www.ooshirts.com/?module=printer&action=orders&view=ReadyToPrint",
]
OO_ACCOUNTS = [
    ("your-email@example.com", "your-password"),        # <-- UPDATE: screen print account
    ("your-dtg-email@example.com", "your-dtg-password"), # <-- UPDATE: DTG account
]

# --------- logging ---------
logger = logging.getLogger("ooshirts_only")
logger.setLevel(logging.INFO)
if not logger.handlers:
    fh = RotatingFileHandler(str(LOG_PATH), maxBytes=1_000_000, backupCount=2, encoding="utf-8")
    fh.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
    logger.addHandler(fh)
    sh = logging.StreamHandler()
    sh.setFormatter(logging.Formatter("%(message)s"))
    logger.addHandler(sh)

def log_ex(prefix="error"):
    try:
        logger.exception(prefix)
    except Exception:
        print(prefix)
        traceback.print_exc()

# --------- date helpers ---------
_num_date = re.compile(r'\b(0?[1-9]|1[0-2])/(0?[1-9]|[12]\d|3[01])(?:/\d{2,4})?\b')
_wk_mmm_dd = re.compile(r'\b(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)\.?\s+([0-3]?\d)\b', re.I)
_mmm_dd    = re.compile(r'\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)\.?\s+([0-3]?\d)\b', re.I)

def to_mmdd(s: str) -> str:
    """Normalize various 'Ship By' formats to MM/DD; ignore things like '0/1 Received'."""
    s = (s or "").strip()
    # Try: Tue Aug 12
    m = _wk_mmm_dd.search(s)
    if m:
        month_map = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,'jul':7,'aug':8,'sep':9,'sept':9,'oct':10,'nov':11,'dec':12}
        mm = month_map[m.group(1).lower()]
        dd = int(m.group(2))
        return f"{mm:02d}/{dd:02d}"
    # Try: Aug 12
    m = _mmm_dd.search(s)
    if m:
        month_map = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,'jul':7,'aug':8,'sep':9,'sept':9,'oct':10,'nov':11,'dec':12}
        mm = month_map[m.group(1).lower()]
        dd = int(m.group(2))
        return f"{mm:02d}/{dd:02d}"
    # Try: 08/12
    m = _num_date.search(s)
    if m and "received" not in s.lower():   # guard against "0/1 Received"
        seg = m.group(0)
        if seg.count("/") == 1:
            seg = f"{seg}/{datetime.now().year}"
        try:
            dt = datetime.strptime(seg, "%m/%d/%Y")
            return dt.strftime("%m/%d")
        except Exception:
            pass
    return ""

# --------- workbook helpers ---------
HEADER = ["Order #", "Status", "Vendor", "Units", "Screens", "Shipper", "Check-In", "Due Date"]

def ensure_workbook(path: Path):
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        # ensure header and Due Date column
        if ws.max_row == 0 or (ws.cell(row=1, column=1).value or "").strip() != "Order #":
            ws.insert_rows(1)
            for i, h in enumerate(HEADER, start=1):
                ws.cell(row=1, column=i, value=h)
        else:
            # add Due Date if missing
            existing = [str(ws.cell(row=1, column=i).value or "") for i in range(1, ws.max_column+1)]
            if "Due Date" not in existing:
                ws.cell(row=1, column=ws.max_column+1, value="Due Date")
    else:
        wb = Workbook()
        ws = wb.active
        for i, h in enumerate(HEADER, start=1):
            ws.cell(row=1, column=i, value=h)
    return wb, ws

def upsert_rows(path: Path, rows):
    wb, ws = ensure_workbook(path)
    index = {}
    for i in range(2, ws.max_row + 1):
        key = str(ws.cell(row=i, column=1).value or "").strip()
        if key:
            index[key] = i
    added, replaced = 0, 0
    for r in rows:
        if not r or not r[0]:
            continue
        k = str(r[0]).strip()
        if k in index:
            ridx = index[k]
            for c in range(1, len(HEADER)+1):
                ws.cell(row=ridx, column=c, value=r[c-1] if c-1 < len(r) else "")
            replaced += 1
        else:
            ws.append(r); added += 1
            index[k] = ws.max_row
    try:
        wb.save(path)
        logger.info(f"[save] added {added}, replaced {replaced} -> {path}")
    except PermissionError:
        alt = path.with_name(path.stem + "_NEW.xlsx")
        wb.save(alt)
        logger.warning(f"[save] workbook is open; saved to {alt}")

# --------- scraping ---------
def login_workflow(driver, email, password):
    """Simple login used in your old script."""
    DASH = "https://www.ooshirts.com/index.php?module=printer&action=dashboard"
    driver.get(DASH)
    try:
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.NAME, "email")))
    except Exception:
        driver.get("https://www.ooshirts.com/?module=printer&action=orders&view=ReadyToPrint")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.NAME, "email")))
    driver.find_element(By.NAME, "email").clear()
    driver.find_element(By.NAME, "email").send_keys(email)
    driver.find_element(By.NAME, "password").clear()
    driver.find_element(By.NAME, "password").send_keys(password)
    try:
        driver.find_element(By.XPATH, "//input[@type='submit']").click()
    except Exception:
        driver.find_element(By.NAME, "password").send_keys(Keys.ENTER)
    WebDriverWait(driver, 20).until(lambda d: "action=login" not in d.current_url)
    logger.info(f"[login] {email} -> {driver.current_url}")

def find_orders_table(drv):
    """Find the Orders table that contains the 'Ship By' header."""
    tables = drv.find_elements(By.TAG_NAME, "table")
    for t in tables:
        try:
            headers = [th.text.strip() for th in t.find_elements(By.TAG_NAME, "th")]
            low = [h.lower() for h in headers]
            if any("ship" in h and "by" in h for h in low) and any("id" == h or h.endswith("id") or "id" in h for h in low):
                return t, headers
        except Exception:
            continue
    return None, []

def extract_orders_from_table(drv, vendor_label):
    """Strictly pull ID + Ship By from the Orders table; return rows ready for workbook."""
    table, headers = find_orders_table(drv)
    if not table:
        logger.warning("[parse] Orders table not found (no 'Ship By' header).")
        return []

    low = [h.lower() for h in headers]
    # resolve indices
    id_idx = next(i for i,h in enumerate(low) if h == "id" or h.endswith("id") or "id" in h)
    ship_idx = next(i for i,h in enumerate(low) if "ship" in h and "by" in h)

    # rows
    rows = table.find_elements(By.CSS_SELECTOR, "tbody tr")
    if not rows:
        rows = table.find_elements(By.XPATH, ".//tr[td]")

    out = []
    debug_preview = []
    for tr in rows:
        tds = tr.find_elements(By.TAG_NAME, "td")
        if len(tds) <= max(id_idx, ship_idx):
            continue

        # Order #
        order_raw = tds[id_idx].text.strip()
        try:
            a = tds[id_idx].find_element(By.TAG_NAME, "a")
            if a.text.strip():
                order_raw = a.text.strip()
        except Exception:
            pass
        order = "".join(ch for ch in order_raw if ch.isdigit())
        if not order:
            continue

        # Ship By -> MM/DD
        ship_text = tds[ship_idx].text.strip()
        mmdd = to_mmdd(ship_text)
        debug_preview.append((order_raw, ship_text, mmdd))
        if not mmdd:
            # do not fallback to other cells; skip to avoid '0/1 Received'
            continue

        out.append([order, "", vendor_label, "", "", "", "", mmdd])

    # Log first 8 mappings for sanity
    for row in debug_preview[:8]:
        logger.info(f"[peek] ID cell='{row[0]}' | ShipBy cell='{row[1]}' -> {row[2]}")

    logger.info(f"[parse] extracted {len(out)} rows from table.")
    return out

def scrape_ooshirts():
    all_rows = []
    for email, password in OO_ACCOUNTS:
        vendor = "ooshirts DTG" if "+dtg" in email.lower() else "ooshirts"
        # HEADLESS MODE - runs in background without visible browser
        opts = Options()
        opts.add_argument("--headless=new")  # New headless mode (Chrome 109+)
        opts.add_argument("--disable-gpu")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--window-size=1920,1080")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
        
        drv = webdriver.Chrome(options=opts, service=Service())
        try:
            login_workflow(drv, email, password)
            for url in OO_VIEWS:
                drv.get(url)
                time.sleep(1.5)
                try:
                    body = drv.find_element(By.TAG_NAME, "body")
                    for _ in range(3):
                        body.send_keys(Keys.END); time.sleep(0.2)
                except Exception:
                    pass
                rows = extract_orders_from_table(drv, vendor)
                logger.info(f"[{vendor}] {url.split('view=')[-1]}: {len(rows)} rows")
                all_rows.extend(rows)
        except Exception:
            log_ex(f"[{vendor}] scrape error")
        finally:
            try: drv.quit()
            except Exception: pass
    return all_rows

# --------- main ---------
HEADER = ["Order #", "Status", "Vendor", "Units", "Screens", "Shipper", "Check-In", "Due Date"]

def main():
    logger.info("=" * 50)
    logger.info("OOSHIRTS SCRAPER STARTED (HEADLESS MODE)")
    logger.info("=" * 50)
    
    rows = scrape_ooshirts()
    if not rows:
        logger.info("No Ooshirts rows found.")
        return
    upsert_rows(EXCEL_PATH, rows)

if __name__ == "__main__":
    try:
        main()
        logger.info("[OK] Ooshirts Scraper completed successfully")
        print("\n[OK] Done. Output saved to:", EXCEL_PATH)
        print("(Log:", LOG_PATH, ")")
    except Exception as e:
        error_msg = traceback.format_exc()
        log_ex("Fatal error")
        # Send email alert
        send_error_email("Scraper Failed", error_msg)
        print("\n[!] An error occurred. See ooshirts_scraper.log for details.")
